import csv
import logging
from typing import List, Dict
from openpyxl import load_workbook

logger = logging.getLogger(__name__)

class ParserFactory:
    @staticmethod
    def parse(file_path: str, file_type: str) -> List[Dict]:
        if file_type == "csv":
            return ParserFactory._parse_csv(file_path)
        elif file_type == "xlsx":
            return ParserFactory._parse_xlsx(file_path)
        else:
            raise ValueError(f"Unsupported file type: {file_type}")

    @staticmethod
    def _parse_csv(file_path: str) -> List[Dict]:
        data = []
        with open(file_path, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                data.append(row)
        return data

    @staticmethod
    def _parse_xlsx(file_path: str) -> List[Dict]:
        data = []
        wb = load_workbook(file_path)
        ws = wb.active
        
        headers = [cell.value for cell in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            data.append(dict(zip(headers, row)))
        
        return data