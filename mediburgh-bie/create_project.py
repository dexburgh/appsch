import os

files = {
    "requirements.txt": """fastapi==0.104.1
uvicorn==0.24.0
pydantic==2.5.0
pydantic-settings==2.1.0
python-multipart==0.0.6
openpyxl==3.11.0
pandas==2.1.3
python-dotenv==1.0.0
pytest==7.4.3
pytest-asyncio==0.21.1
httpx==0.25.2""",

    ".env.example": """HOST=127.0.0.1
PORT=8000
RELOAD=True
MAX_FILE_SIZE_MB=50
UPLOAD_DIR=./uploads
ALLOWED_FORMATS=csv,xlsx,pdf
LOG_LEVEL=INFO
DATABASE_URL=sqlite:///./bie.db""",

    ".gitignore": """__pycache__/
*.py[cod]
.env
.venv
venv/
uploads/
*.db""",

    "run.py": """import uvicorn
import os
from dotenv import load_dotenv

load_dotenv()

if __name__ == "__main__":
    host = os.getenv("HOST", "127.0.0.1")
    port = int(os.getenv("PORT", 8000))
    
    uvicorn.run(
        "app.main:app",
        host=host,
        port=port,
        reload=True
    )""",

    "app/__init__.py": "",
    "app/config.py": """import os
from pathlib import Path
from pydantic_settings import BaseSettings

class Settings(BaseSettings):
    HOST: str = os.getenv("HOST", "127.0.0.1")
    PORT: int = int(os.getenv("PORT", 8000))
    MAX_FILE_SIZE_MB: int = int(os.getenv("MAX_FILE_SIZE_MB", 50))
    UPLOAD_DIR: str = os.getenv("UPLOAD_DIR", "./uploads")
    LOG_LEVEL: str = os.getenv("LOG_LEVEL", "INFO")

settings = Settings()
Path(settings.UPLOAD_DIR).mkdir(parents=True, exist_ok=True)""",

    "app/schemas/__init__.py": """from enum import Enum
from typing import Optional, List
from pydantic import BaseModel, Field

class ValidationStatus(str, Enum):
    PENDING = "pending"
    VALID = "valid"
    WARNING = "warning"
    ERROR = "error"

class BillingLineItem(BaseModel):
    claim_id: str
    patient_id: str
    service_date: str
    cpt_code: str
    cpt_code_normalized: Optional[str] = None
    icd10_primary: str
    icd10_primary_normalized: Optional[str] = None
    charge_amount: float
    validation_status: ValidationStatus = ValidationStatus.PENDING

class BillingClaimResponse(BaseModel):
    file_name: str
    file_type: str
    rows_processed: int
    normalized_data: List[BillingLineItem] = []
    total_charges: float = 0.0
    processing_time_seconds: float
    errors: List[str] = []""",

    "app/ingestion/__init__.py": "",
    "app/ingestion/file_handler.py": """import os
import logging
from pathlib import Path
from fastapi import UploadFile

logger = logging.getLogger(__name__)

class FileHandler:
    ALLOWED_FORMATS = ["csv", "xlsx"]
    MAX_FILE_SIZE_BYTES = 50 * 1024 * 1024

    @staticmethod
    def validate_file(file: UploadFile):
        if not file.filename:
            raise ValueError("No file provided")
        
        ext = Path(file.filename).suffix.lower().lstrip(".")
        if ext not in FileHandler.ALLOWED_FORMATS:
            raise ValueError(f"Invalid format. Use CSV or XLSX")

    @staticmethod
    async def save_file(file: UploadFile, upload_dir: str) -> str:
        Path(upload_dir).mkdir(parents=True, exist_ok=True)
        file_path = os.path.join(upload_dir, file.filename)
        
        content = await file.read()
        if len(content) > FileHandler.MAX_FILE_SIZE_BYTES:
            raise ValueError("File too large")
        
        with open(file_path, "wb") as f:
            f.write(content)
        
        return file_path

    @staticmethod
    def get_file_type(file_path: str) -> str:
        return Path(file_path).suffix.lower().lstrip(".")""",

    "app/parsing/__init__.py": "",
    "app/parsing/parser_factory.py": """import csv
import logging
from typing import List, Dict
from openpyxl import load_workbook

logger = logging.getLogger(__name__)

class ParserFactory:
    @staticmethod
    def parse(file_path: str, file_type: str) -> List[Dict]:
        if file_type == "csv":
            return ParserFactory._parse_csv(file_path)
        elif file_type == "xlsx":
            return ParserFactory._parse_xlsx(file_path)
        else:
            raise ValueError(f"Unsupported file type: {file_type}")

    @staticmethod
    def _parse_csv(file_path: str) -> List[Dict]:
        data = []
        with open(file_path, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                data.append(row)
        return data

    @staticmethod
    def _parse_xlsx(file_path: str) -> List[Dict]:
        data = []
        wb = load_workbook(file_path)
        ws = wb.active
        
        headers = [cell.value for cell in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            data.append(dict(zip(headers, row)))
        
        return data""",

    "app/normalization/__init__.py": "",
    "app/normalization/normalizers.py": """import re

class CPTNormalizer:
    @staticmethod
    def normalize(code: str) -> str:
        return str(code).strip().upper().replace("-", "")

class ICD10Normalizer:
    @staticmethod
    def normalize(code: str) -> str:
        code = str(code).strip().upper()
        return code

class NormalizationPipeline:
    def __init__(self):
        self.cpt = CPTNormalizer()
        self.icd10 = ICD10Normalizer()

    def normalize_record(self, raw_record: dict):
        return {
            "claim_id": raw_record.get("claim_id", ""),
            "patient_id": raw_record.get("patient_id", ""),
            "service_date": str(raw_record.get("service_date", "")),
            "cpt_code": raw_record.get("cpt_code", ""),
            "cpt_code_normalized": self.cpt.normalize(raw_record.get("cpt_code", "")),
            "icd10_primary": raw_record.get("icd10_primary", ""),
            "icd10_primary_normalized": self.icd10.normalize(raw_record.get("icd10_primary", "")),
            "charge_amount": float(raw_record.get("charge_amount", 0)),
        }

    def normalize_batch(self, records: list) -> tuple:
        normalized = []
        errors = []
        
        for idx, record in enumerate(records):
            try:
                normalized.append(self.normalize_record(record))
            except Exception as e:
                errors.append(f"Row {idx + 1}: {str(e)}")
        
        return normalized, errors""",

    "app/validation/__init__.py": "",
    "app/validation/validators.py": """from app.schemas import ValidationStatus

class ValidationManager:
    def validate_item(self, item):
        item["validation_status"] = ValidationStatus.PENDING
        return item""",

    "app/export/__init__.py": "",
    "app/export/exporter.py": """import json
import csv
from pathlib import Path

class Exporter:
    @staticmethod
    def to_json(data, output_path: str) -> str:
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        with open(output_path, 'w') as f:
            json.dump(data, f, indent=2)
        return output_path

    @staticmethod
    def to_csv(data, output_path: str) -> str:
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        if not data:
            return output_path
        
        with open(output_path, 'w', newline='') as f:
            writer = csv.DictWriter(f, fieldnames=data[0].keys())
            writer.writeheader()
            writer.writerows(data)
        
        return output_path""",

    "app/main.py": """import logging
import time
from datetime import datetime
from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.middleware.cors import CORSMiddleware

from app.config import settings
from app.ingestion.file_handler import FileHandler
from app.parsing.parser_factory import ParserFactory
from app.normalization.normalizers import NormalizationPipeline
from app.validation.validators import ValidationManager
from app.schemas import BillingClaimResponse, HealthCheckResponse, ValidationStatus

logging.basicConfig(level=settings.LOG_LEVEL)
logger = logging.getLogger(__name__)

app = FastAPI(title="MediBurgh BIE", version="1.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

pipeline = NormalizationPipeline()
validator = ValidationManager()

@app.get("/")
async def root():
    return {"name": "MediBurgh BIE", "status": "operational"}

@app.get("/health")
async def health_check():
    return HealthCheckResponse(
        status="healthy",
        version="1.0.0",
        timestamp=datetime.utcnow().isoformat()
    )

@app.post("/api/v1/billing/upload")
async def upload_billing_file(file: UploadFile = File(...)):
    start = time.time()
    
    try:
        FileHandler.validate_file(file)
        file_path = await FileHandler.save_file(file, settings.UPLOAD_DIR)
        file_type = FileHandler.get_file_type(file_path)
        
        raw_data = ParserFactory.parse(file_path, file_type)
        if not raw_data:
            raise ValueError("No data found")
        
        normalized, errors = pipeline.normalize_batch(raw_data)
        
        total = sum(item.get("charge_amount", 0) for item in normalized)
        
        return BillingClaimResponse(
            file_name=file.filename,
            file_type=file_type,
            rows_processed=len(raw_data),
            normalized_data=normalized,
            total_charges=total,
            processing_time_seconds=round(time.time() - start, 3),
            errors=errors
        )
    
    except Exception as e:
        logger.error(f"Error: {str(e)}")
        raise HTTPException(status_code=400, detail=str(e))""",

    "tests/__init__.py": "",
    "tests/test_main.py": """from fastapi.testclient import TestClient
from app.main import app

client = TestClient(app)

def test_health():
    response = client.get("/health")
    assert response.status_code == 200
    assert response.json()["status"] == "healthy\"""",

    "docs/API.md": """# API

## POST /api/v1/billing/upload
Upload CSV or XLSX file

## GET /health
Health check""",

    "samples/sample_billing.csv": """claim_id,patient_id,service_date,cpt_code,icd10_primary,charge_amount
CLM-001,PAT-001,2024-01-15,99213,I10,150.00
CLM-002,PAT-002,2024-01-16,99214,J44.9,200.00"""
}

def create():
    for path, content in files.items():
        os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
        with open(path, 'w', encoding='utf-8') as f:
            f.write(content)
        print(f"✓ {path}")

if __name__ == "__main__":
    create()
    print("\n✅ Done!")